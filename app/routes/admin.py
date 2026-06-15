import os
import csv
import io
import json
import secrets
import smtplib
from email.message import EmailMessage
from datetime import datetime, timedelta
from collections import OrderedDict
from flask import Blueprint, render_template, request, redirect, url_for, flash, abort, send_file, current_app, Response
from flask_login import login_required, current_user
from ..security import require_roles, validate_password_strength, is_unit_management_account, normalize_unit_appointment, unit_appointment_label, is_unit_ao_account, is_unit_workspace_account
from .. import db
from ..models import User, Unit, UnitLevel, Role, Channel, Settings, AuditEvent, Broadcast, BroadcastReceipt, BroadcastAck, UserSession
from ..audit import log_event
from ..backup import build_backup_zip_bytes, backup_folder

bp = Blueprint("admin", __name__, url_prefix="/admin")

RANK_OPTIONS = [
    "ACM", "LAC", "CPL", "SGT", "FS", "WO", "MWO", "AWO",
    "PLT OFFR", "FG OFFR", "FLT LT", "SQN LDR", "WG CDR", "GP CAPT",
    "AIR CDRE", "AVM", "AIR MSHL", "AIR CHIEF MSHL"
]

ROLE_OPTIONS = [
    (Role.OFFICER.value, "View-Only Officer"),
    (Role.COMMANDER.value, "Unit Commander"),
    (Role.ADMIN.value, "General Admin (HQ)"),
]

UNIT_APPOINTMENT_OPTIONS = [
    ("CHIEF_CLERK", "Chief Clerk"),
    ("ADMIN_OFFICER", "Admin Officer / Unit AO"),
    ("SIGNATORY_OFFICER", "Signatory Officer"),
    ("OTHER_OFFICER", "Other Officer"),
    ("COMMANDER", "Commander"),
]

def _can_manage_unit(unit_id: int | None = None) -> bool:
    if not getattr(current_user, "is_authenticated", False):
        return False
    if getattr(current_user, "role", "") in (Role.ADMIN.value, Role.SUPER_ADMIN.value):
        return True
    if not is_unit_management_account(current_user):
        return False
    if unit_id is None:
        return bool(getattr(current_user, "unit_id", None))
    return int(unit_id) == int(getattr(current_user, "unit_id", 0) or 0)

def _unit_scope_or_403(unit_id=None):
    if not _can_manage_unit(unit_id):
        abort(403)
    if getattr(current_user, "role", "") in (Role.ADMIN.value, Role.SUPER_ADMIN.value):
        return int(unit_id) if unit_id else None
    return int(current_user.unit_id)

def _role_for_unit_appointment(appointment_key: str) -> str:
    # Keep HQ-level ADMIN separate from unit AO. Unit permissions are controlled
    # by appointment, not by giving every unit AO global admin rights.
    if appointment_key == "COMMANDER":
        return Role.COMMANDER.value
    return Role.MEMBER.value

def _unit_primary_ao(unit_id: int | None):
    if not unit_id:
        return None
    return (User.query.filter(User.unit_id == unit_id, User.is_active_flag == True)
            .filter(User.appointment.in_(["Admin Officer", "Admin Officer / Unit AO", "Unit AO"]))
            .order_by(User.created_at.asc()).first())

def _unit_admin_summary(unit_id: int):
    officers = User.query.filter(User.unit_id == unit_id).all()
    def appt_count(label):
        return len([u for u in officers if normalize_unit_appointment(u.appointment) == label])
    return {
        "total": len(officers),
        "active": len([u for u in officers if u.is_active_flag]),
        "chief_clerks": appt_count("CHIEF_CLERK"),
        "aos": appt_count("ADMIN_OFFICER"),
        "signatories": appt_count("SIGNATORY_OFFICER"),
        "commanders": appt_count("COMMANDER"),
    }


def _unit_account_for(unit_id: int | None):
    if not unit_id:
        return None
    return User.query.filter(User.unit_id == unit_id, User.account_type == "UNIT").order_by(User.created_at.asc()).first()

def _unit_account_service_number(unit: Unit) -> str:
    return f"UNIT/{(unit.code or '').strip().upper()}"

def _global_user_query():
    """Global user directory excludes unit workspace accounts and internal unit officers.

    The corrected structure is:
    - /admin/users: HQ/global accounts and view-only officers only.
    - /admin/unit-accounts: workspace login linked to an existing Unit.
    - /admin/unit-officers: Chief Clerk, AO, Commander, Signatory and Other Officers inside a unit.
    """
    unit_appts = {unit_appointment_label(k) for k, _ in UNIT_APPOINTMENT_OPTIONS}
    return (User.query
            .filter(User.account_type != "UNIT")
            .filter(User.role.in_([Role.SUPER_ADMIN.value, Role.ADMIN.value, Role.OFFICER.value]))
            .filter((User.appointment.is_(None)) | (~User.appointment.in_(list(unit_appts)))))

def _create_unit_workspace_account(unit: Unit, email: str | None, password: str | None):
    service_number = _unit_account_service_number(unit)
    existing = User.query.filter((User.service_number == service_number) | ((User.unit_id == unit.id) & (User.account_type == "UNIT"))).first()
    if existing:
        return None, f"{unit.code} already has a unit workspace account."
    temp_password = (password or "").strip() or _suggest_export_password("UNIT")
    acct = User(
        full_name=f"{unit.name} Unit Workspace",
        service_number=service_number,
        rank=None,
        email=(email or "").strip().lower() or None,
        appointment="Unit Account",
        account_type="UNIT",
        role=Role.MEMBER.value,
        unit_id=unit.id,
        clearance_level="RESTRICTED",
        must_change_password=True,
        password_hash="tmp",
    )
    acct.set_password(temp_password)
    db.session.add(acct)
    db.session.commit()
    log_event(current_user.id, "UNIT_WORKSPACE_ACCOUNT_CREATED", {"unit_id": unit.id, "unit": unit.code, "service_number": service_number})
    return acct, temp_password

CLEARANCE_BY_RANK = {
    "ACM": "RESTRICTED",
    "LAC": "RESTRICTED",
    "CPL": "RESTRICTED",
    "SGT": "RESTRICTED",
    "FS": "RESTRICTED",
    "WO": "CONFIDENTIAL",
    "MWO": "CONFIDENTIAL",
    "AWO": "CONFIDENTIAL",
    "PLT OFFR": "RESTRICTED",
    "FG OFFR": "RESTRICTED",
    "FLT LT": "CONFIDENTIAL",
    "SQN LDR": "SECRET",
    "WG CDR": "TOP SECRET",
    "GP CAPT": "TOP SECRET",
    "AIR CDRE": "TOP SECRET",
    "AVM": "TOP SECRET",
    "AIR MSHL": "TOP SECRET",
    "AIR CHIEF MSHL": "TOP SECRET",
}



def _suggest_export_password(prefix="NMS") -> str:
    alphabet = "ABCDEFGHJKLMNPQRSTUVWXYZ23456789"
    tail = ''.join(secrets.choice(alphabet) for _ in range(10))
    return f"{prefix}-{datetime.utcnow().strftime('%Y%m%d')}-{tail}"

def _resolve_export_password():
    mode = (request.form.get("password_mode") or "manual").strip().lower()
    if mode == "auto":
        password = (request.form.get("suggested_password") or "").strip()
    else:
        password = (request.form.get("export_password") or "").strip()
    if not password:
        return None, "Export password is required."
    if len(password) < 8:
        return None, "Export password must be at least 8 characters."
    return password, None

def _password_protect_zip_bytes(source_bytes: bytes, inner_name: str, password: str, comment: str = "") -> bytes:
    try:
        import pyzipper
    except Exception as exc:
        raise RuntimeError("pyzipper is required for password-protected ZIP exports. Install it with: pip install pyzipper") from exc
    mem = io.BytesIO()
    with pyzipper.AESZipFile(mem, "w", compression=pyzipper.ZIP_DEFLATED, encryption=pyzipper.WZ_AES) as zf:
        zf.setpassword(password.encode("utf-8"))
        if comment:
            zf.comment = comment.encode("utf-8")[:65535]
        zf.writestr(inner_name, source_bytes)
    return mem.getvalue()

def _safe_role(role: str) -> str:
    allowed = {r[0] for r in ROLE_OPTIONS}
    if role == Role.SUPER_ADMIN.value and current_user.role == Role.SUPER_ADMIN.value:
        return role
    return role if role in allowed else Role.OFFICER.value

def _rank_clearance(rank: str | None, fallback="RESTRICTED") -> str:
    return CLEARANCE_BY_RANK.get((rank or "").strip(), fallback)

def _send_login_email(user: User, temporary_password: str) -> bool:
    if not getattr(user, "email", None):
        return False
    host = os.getenv("SMTP_HOST")
    sender = os.getenv("SMTP_FROM") or os.getenv("SMTP_USER")
    if not host or not sender:
        return False
    msg = EmailMessage()
    msg["Subject"] = "NAF Secure Messaging Platform Login Details"
    msg["From"] = sender
    msg["To"] = user.email
    msg.set_content(f"""Dear {user.full_name},

Your NAF Secure Messaging Platform account has been created.

Username/Service Number: {user.service_number}
Temporary Password: {temporary_password}
Role: {user.role}
Clearance: {user.clearance_level}

You will be required to change this temporary password after first login.

Regards,
HQ Admin
""")
    port = int(os.getenv("SMTP_PORT", "587"))
    username = os.getenv("SMTP_USER")
    password = os.getenv("SMTP_PASSWORD")
    try:
        with smtplib.SMTP(host, port, timeout=10) as smtp:
            if os.getenv("SMTP_TLS", "1").lower() in {"1", "true", "yes", "on"}:
                smtp.starttls()
            if username and password:
                smtp.login(username, password)
            smtp.send_message(msg)
        return True
    except Exception:
        current_app.logger.exception("Failed to send onboarding email to %s", user.email)
        return False



# =========================
# Bulk import helpers
# =========================
def _normalize_header(value):
    return (str(value or "").strip().lower().replace(" ", "_").replace("-", "_"))

def _read_import_rows(upload):
    """Read CSV/XLSX upload into a list of dictionaries with normalized headers."""
    filename = (upload.filename or "").lower()
    raw = upload.read()
    rows = []
    if filename.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise RuntimeError("XLSX import requires openpyxl. Install it with: pip install openpyxl") from exc
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        data = list(ws.iter_rows(values_only=True))
        if not data:
            return []
        headers = [_normalize_header(h) for h in data[0]]
        for row in data[1:]:
            item = {headers[i]: ("" if row[i] is None else str(row[i]).strip()) for i in range(min(len(headers), len(row))) if headers[i]}
            if any(v for v in item.values()):
                rows.append(item)
    else:
        text = raw.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            item = {_normalize_header(k): (v or "").strip() for k, v in row.items() if k}
            if any(v for v in item.values()):
                rows.append(item)
    return rows

def _row_value(row, *names):
    for name in names:
        val = row.get(_normalize_header(name))
        if val not in (None, ""):
            return str(val).strip()
    return ""

def _import_report_response(title, rows, filename):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Import", title])
    writer.writerow([])
    writer.writerow(["Row", "Status", "Identity", "Message"])
    for r in rows:
        writer.writerow([r.get("row"), r.get("status"), r.get("identity"), r.get("message")])
    return Response(output.getvalue(), mimetype="text/csv", headers={"Content-Disposition": f"attachment; filename={filename}"})

@bp.get("/")
@login_required
@require_roles(Role.ADMIN.value)
def admin_home():
    users = User.query.count()
    units = Unit.query.count()
    channels = Channel.query.count()
    settings = Settings.get()
    audits = []
    if current_user.role == Role.SUPER_ADMIN.value:
        audits = AuditEvent.query.order_by(AuditEvent.created_at.desc()).limit(15).all()
    return render_template("admin/index.html", users=users, units=units, channels=channels, audits=audits, settings=settings)



@bp.get("/users/import-template.csv")
@login_required
@require_roles(Role.ADMIN.value)
def users_import_template():
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["service_number", "full_name", "rank", "email", "unit_code", "appointment", "specialty", "phone", "clearance_level"])
    writer.writerow(["NAF/0001", "FLT LT SAMPLE OFFICER", "Flight Lieutenant", "sample.officer@naf.local", "641CISG", "Drafter", "Communications", "08000000000", "CONFIDENTIAL"])
    return Response(output.getvalue(), mimetype="text/csv", headers={"Content-Disposition": "attachment; filename=nms_officers_import_template.csv"})

@bp.post("/users/import")
@login_required
@require_roles(Role.ADMIN.value)
def import_users():
    upload = request.files.get("import_file")
    if not upload or not upload.filename:
        flash("Select a CSV or XLSX file to import officers.", "warning")
        return redirect(url_for("admin.users"))
    try:
        rows = _read_import_rows(upload)
    except Exception as exc:
        flash(str(exc), "danger")
        return redirect(url_for("admin.users"))
    report = []
    created = skipped = failed = 0
    units_by_code = {u.code.upper(): u for u in Unit.query.all()}
    for idx, row in enumerate(rows, start=2):
        service_number = _row_value(row, "service_number", "service no", "svc_no", "number").upper()
        full_name = _row_value(row, "full_name", "name")
        rank = _row_value(row, "rank") or None
        email = _row_value(row, "email", "official_email").lower()
        unit_code = _row_value(row, "unit_code", "unit").upper()
        appointment = _row_value(row, "appointment") or None
        specialty = _row_value(row, "specialty", "trade") or None
        phone = _row_value(row, "phone", "phone_number") or None
        clearance_level = (_row_value(row, "clearance_level", "clearance") or _rank_clearance(rank)).upper()
        identity = service_number or email or full_name or f"row {idx}"
        if not service_number or not full_name or not email:
            failed += 1; report.append({"row": idx, "status": "FAILED", "identity": identity, "message": "service_number, full_name and email are required"}); continue
        if User.query.filter_by(service_number=service_number).first():
            skipped += 1; report.append({"row": idx, "status": "SKIPPED", "identity": identity, "message": "service_number already exists"}); continue
        if User.query.filter(User.email.ilike(email)).first():
            skipped += 1; report.append({"row": idx, "status": "SKIPPED", "identity": identity, "message": "email already exists"}); continue
        unit = units_by_code.get(unit_code) if unit_code else None
        u = User(full_name=full_name, service_number=service_number, rank=rank, specialty=specialty, email=email, phone=phone, appointment=appointment, account_type="OFFICER", role=Role.OFFICER.value, unit_id=unit.id if unit else None, clearance_level=clearance_level, must_change_password=True, password_hash="tmp")
        u.set_password(email)
        db.session.add(u)
        created += 1; report.append({"row": idx, "status": "CREATED", "identity": identity, "message": "view-only officer created; initial password is official email"})
    db.session.commit()
    log_event(current_user.id, "BULK_OFFICER_IMPORT", {"created": created, "skipped": skipped, "failed": failed, "status": "SUCCESS"})
    flash(f"Officer import complete: {created} created, {skipped} skipped, {failed} failed. Report downloaded.", "success" if failed == 0 else "warning")
    return _import_report_response("Officer Bulk Import", report, "nms_officer_import_report.csv")

@bp.get("/unit-admin")
@login_required
def unit_admin_center():
    """Phase 2: Unit AO command/admin center.

    This is the unit-level super-admin landing page. It keeps the unit account
    design clear: AO manages unit users and oversight inside only one unit; HQ
    operators can inspect any unit.
    """
    _unit_scope_or_403()
    selectable_units = Unit.query.order_by(Unit.name.asc()).all() if current_user.role in (Role.ADMIN.value, Role.SUPER_ADMIN.value) else []
    managed_unit_id = request.args.get("unit_id", type=int) if current_user.role in (Role.ADMIN.value, Role.SUPER_ADMIN.value) else current_user.unit_id
    if not managed_unit_id and selectable_units:
        managed_unit_id = selectable_units[0].id
    managed_unit_id = _unit_scope_or_403(managed_unit_id)
    unit = Unit.query.get_or_404(managed_unit_id)
    officers = User.query.filter(User.unit_id == unit.id).order_by(User.appointment.asc(), User.full_name.asc()).all()
    primary_ao = _unit_primary_ao(unit.id)
    summary = _unit_admin_summary(unit.id)
    recent_audit = (AuditEvent.query.join(User, AuditEvent.actor_id == User.id)
                    .filter(User.unit_id == unit.id)
                    .order_by(AuditEvent.created_at.desc()).limit(30).all())
    recent_signals = (Broadcast.query.filter(Broadcast.from_unit_id == unit.id)
                      .order_by(Broadcast.created_at.desc()).limit(12).all())
    log_event(current_user.id, "UNIT_ADMIN_CENTER_VIEWED", {"unit_id": unit.id, "unit": unit.code})
    return render_template(
        "admin/unit_admin_center.html",
        unit=unit,
        units=selectable_units,
        officers=officers,
        primary_ao=primary_ao,
        summary=summary,
        recent_audit=recent_audit,
        recent_signals=recent_signals,
        is_current_ao=is_unit_ao_account(current_user),
        is_workspace=is_unit_workspace_account(current_user),
    )

@bp.route("/unit-officers", methods=["GET", "POST"])
@login_required
def unit_officers():
    """Unit-level personnel management foundation.

    Phase 1: a unit account, Unit Commander, or Unit AO can create and manage
    officers inside the same unit. HQ Admin/Super Admin can select any unit.
    """
    _unit_scope_or_403()
    selectable_units = Unit.query.order_by(Unit.name.asc()).all() if current_user.role in (Role.ADMIN.value, Role.SUPER_ADMIN.value) else []
    managed_unit_id = request.args.get("unit_id", type=int) if current_user.role in (Role.ADMIN.value, Role.SUPER_ADMIN.value) else current_user.unit_id
    if not managed_unit_id and selectable_units:
        managed_unit_id = selectable_units[0].id
    managed_unit_id = _unit_scope_or_403(managed_unit_id)
    unit = Unit.query.get_or_404(managed_unit_id)

    if request.method == "POST":
        action = (request.form.get("action") or "create").strip().lower()
        if action == "create":
            full_name = (request.form.get("full_name") or "").strip()
            service_number = (request.form.get("service_number") or "").strip().upper()
            rank = (request.form.get("rank") or "").strip() or None
            email = (request.form.get("email") or "").strip().lower() or None
            phone = (request.form.get("phone") or "").strip() or None
            specialty = (request.form.get("specialty") or "").strip() or None
            appointment_key = normalize_unit_appointment(request.form.get("appointment_key"))
            if appointment_key not in {x[0] for x in UNIT_APPOINTMENT_OPTIONS}:
                appointment_key = "OTHER_OFFICER"
            appointment = unit_appointment_label(appointment_key)
            clearance_level = (request.form.get("clearance_level") or _rank_clearance(rank)).strip().upper()
            if request.form.get("auto_clearance"):
                clearance_level = _rank_clearance(rank, clearance_level)
            password = (request.form.get("password") or "").strip()
            auto_password = False
            if not password:
                password = _suggest_export_password("UNIT")
                auto_password = True
            ok, issues = validate_password_strength(password)
            if not full_name or not service_number:
                flash("Full name and service number are required.", "danger")
                return redirect(url_for("admin.unit_officers", unit_id=unit.id))
            if not ok:
                flash("Password issue: " + "; ".join(issues), "danger")
                return redirect(url_for("admin.unit_officers", unit_id=unit.id))
            if User.query.filter_by(service_number=service_number).first():
                flash("That service number already exists.", "danger")
                return redirect(url_for("admin.unit_officers", unit_id=unit.id))
            if email and User.query.filter(User.email.ilike(email)).first():
                flash("That email address is already assigned to another account.", "danger")
                return redirect(url_for("admin.unit_officers", unit_id=unit.id))
            u = User(
                full_name=full_name,
                service_number=service_number,
                rank=rank,
                specialty=specialty,
                email=email,
                phone=phone,
                appointment=appointment,
                account_type="OFFICER",
                role=_role_for_unit_appointment(appointment_key),
                unit_id=unit.id,
                clearance_level=clearance_level,
                must_change_password=True,
                password_hash="tmp",
            )
            u.set_password(password)
            db.session.add(u)
            db.session.commit()
            sent = _send_login_email(u, password) if request.form.get("email_login_details") else False
            log_event(current_user.id, "UNIT_USER_CREATED", {
                "unit_id": unit.id,
                "unit": unit.code,
                "officer_id": u.id,
                "service_number": u.service_number,
                "appointment": appointment,
                "role": u.role,
            })
            msg = f"{appointment} account created for {u.full_name}."
            if auto_password and not sent:
                msg += f" Temporary password: {password}"
            elif sent:
                msg += " Login details emailed."
            flash(msg, "success")
            return redirect(url_for("admin.unit_officers", unit_id=unit.id))

    officers = User.query.filter(User.unit_id == unit.id).order_by(User.created_at.desc()).all()
    audit = (AuditEvent.query.join(User, AuditEvent.actor_id == User.id)
             .filter(User.unit_id == unit.id)
             .order_by(AuditEvent.created_at.desc()).limit(80).all())
    return render_template(
        "admin/unit_officers.html",
        unit=unit,
        units=selectable_units,
        officers=officers,
        audit=audit,
        rank_options=RANK_OPTIONS,
        appointment_options=UNIT_APPOINTMENT_OPTIONS,
        unit_appointment_label=unit_appointment_label,
        primary_ao=_unit_primary_ao(unit.id),
        summary=_unit_admin_summary(unit.id),
        is_current_ao=is_unit_ao_account(current_user),
        is_workspace=is_unit_workspace_account(current_user),
    )

@bp.route("/unit-officers/<int:user_id>/edit", methods=["GET", "POST"])
@login_required
def edit_unit_officer(user_id):
    u = User.query.get_or_404(user_id)
    _unit_scope_or_403(u.unit_id)
    unit = u.unit
    if request.method == "POST":
        full_name = (request.form.get("full_name") or "").strip()
        service_number = (request.form.get("service_number") or "").strip().upper()
        rank = (request.form.get("rank") or "").strip() or None
        email = (request.form.get("email") or "").strip().lower() or None
        phone = (request.form.get("phone") or "").strip() or None
        specialty = (request.form.get("specialty") or "").strip() or None
        appointment_key = normalize_unit_appointment(request.form.get("appointment_key"))
        if appointment_key not in {x[0] for x in UNIT_APPOINTMENT_OPTIONS}:
            appointment_key = normalize_unit_appointment(u.appointment)
        clearance_level = (request.form.get("clearance_level") or u.clearance_level or "RESTRICTED").strip().upper()
        if not full_name or not service_number:
            flash("Full name and service number are required.", "danger")
            return redirect(url_for("admin.edit_unit_officer", user_id=u.id))
        existing = User.query.filter(User.service_number == service_number, User.id != u.id).first()
        if existing:
            flash("That service number is already assigned to another account.", "danger")
            return redirect(url_for("admin.edit_unit_officer", user_id=u.id))
        if email and User.query.filter(User.email.ilike(email), User.id != u.id).first():
            flash("That email address is already assigned to another account.", "danger")
            return redirect(url_for("admin.edit_unit_officer", user_id=u.id))
        u.full_name = full_name
        u.service_number = service_number
        u.rank = rank
        u.email = email
        u.phone = phone
        u.specialty = specialty
        u.appointment = unit_appointment_label(appointment_key)
        u.role = _role_for_unit_appointment(appointment_key)
        u.clearance_level = clearance_level
        db.session.commit()
        log_event(current_user.id, "UNIT_USER_UPDATED", {
            "officer_id": u.id,
            "service_number": u.service_number,
            "unit_id": u.unit_id,
            "appointment": u.appointment,
        })
        flash("Unit officer updated.", "success")
        return redirect(url_for("admin.unit_officers", unit_id=u.unit_id))
    return render_template("admin/unit_officer_edit.html", u=u, unit=unit, rank_options=RANK_OPTIONS, appointment_options=UNIT_APPOINTMENT_OPTIONS, current_appointment=normalize_unit_appointment(u.appointment))

@bp.post("/unit-officers/<int:user_id>/toggle")
@login_required
def toggle_unit_officer(user_id):
    u = User.query.get_or_404(user_id)
    _unit_scope_or_403(u.unit_id)
    if u.id == current_user.id:
        flash("You cannot disable your own active session account.", "warning")
        return redirect(url_for("admin.unit_officers", unit_id=u.unit_id))
    u.is_active_flag = not bool(u.is_active_flag)
    db.session.commit()
    log_event(current_user.id, "UNIT_USER_STATUS_CHANGED", {
        "officer_id": u.id,
        "service_number": u.service_number,
        "unit_id": u.unit_id,
        "active": u.is_active_flag,
    })
    flash(f"{u.full_name} is now {'active' if u.is_active_flag else 'disabled'}.", "success")
    return redirect(url_for("admin.unit_officers", unit_id=u.unit_id))

@bp.post("/unit-officers/<int:user_id>/reset-password")
@login_required
def reset_unit_officer_password(user_id):
    u = User.query.get_or_404(user_id)
    _unit_scope_or_403(u.unit_id)
    password = (request.form.get("password") or "").strip() or _suggest_export_password("UNIT")
    ok, issues = validate_password_strength(password)
    if not ok:
        flash("Password issue: " + "; ".join(issues), "danger")
        return redirect(url_for("admin.unit_officers", unit_id=u.unit_id))
    u.set_password(password)
    u.must_change_password = True
    db.session.commit()
    sent = _send_login_email(u, password)
    log_event(current_user.id, "UNIT_USER_PASSWORD_RESET", {
        "target_user_id": u.id,
        "service_number": u.service_number,
        "unit_id": u.unit_id,
        "reset_by": current_user.id,
        "email_sent": sent,
        "force_change": True,
    })
    flash(("Password updated and emailed." if sent else f"Password updated. Temporary password: {password}"), "success")
    return redirect(url_for("admin.unit_officers", unit_id=u.unit_id))

@bp.get("/units/import-template.csv")
@login_required
@require_roles(Role.ADMIN.value)
def units_import_template():
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["unit_code", "unit_name", "level", "parent_code"])
    writer.writerow(["641CISG", "641 Communication and Information Systems Group", "UNIT", ""])
    return Response(output.getvalue(), mimetype="text/csv", headers={"Content-Disposition": "attachment; filename=nms_units_import_template.csv"})

@bp.post("/units/import")
@login_required
@require_roles(Role.ADMIN.value)
def import_units():
    upload = request.files.get("import_file")
    if not upload or not upload.filename:
        flash("Select a CSV or XLSX file to import units.", "warning")
        return redirect(url_for("admin.units"))
    try:
        rows = _read_import_rows(upload)
    except Exception as exc:
        flash(str(exc), "danger")
        return redirect(url_for("admin.units"))
    report = []
    created = updated = failed = 0
    # First pass: create/update unit records without parents so parent lookup can work regardless of row order.
    for idx, row in enumerate(rows, start=2):
        code = _row_value(row, "unit_code", "code", "unit").upper()
        name = _row_value(row, "unit_name", "name")
        level = (_row_value(row, "level") or UnitLevel.UNIT.value).upper()
        identity = code or name or f"row {idx}"
        if not code or not name:
            failed += 1; report.append({"row": idx, "status": "FAILED", "identity": identity, "message": "unit_code and unit_name are required"}); continue
        unit = Unit.query.filter_by(code=code).first()
        if unit:
            unit.name = name; unit.level = level; updated += 1
            report.append({"row": idx, "status": "UPDATED", "identity": code, "message": "unit record updated"})
        else:
            unit = Unit(name=name, code=code, level=level)
            db.session.add(unit); created += 1
            report.append({"row": idx, "status": "CREATED", "identity": code, "message": "unit record created"})
    db.session.flush()
    units_by_code = {u.code.upper(): u for u in Unit.query.all()}
    # Second pass: parent hierarchy only. Unit workspace accounts are created
    # separately from /admin/unit-accounts so every account is linked deliberately
    # to an existing unit and not mixed with officer imports.
    for idx, row in enumerate(rows, start=2):
        code = _row_value(row, "unit_code", "code", "unit").upper()
        if not code or code not in units_by_code:
            continue
        unit = units_by_code[code]
        parent_code = _row_value(row, "parent_code", "parent").upper()
        if parent_code and parent_code in units_by_code and parent_code != code:
            unit.parent_id = units_by_code[parent_code].id
    db.session.commit()
    log_event(current_user.id, "BULK_UNIT_IMPORT", {"created": created, "updated": updated, "failed": failed, "status": "SUCCESS"})
    flash(f"Unit import complete: {created} created, {updated} updated, {failed} failed. Create unit workspace accounts from Unit Accounts.", "success" if failed == 0 else "warning")
    return _import_report_response("Unit Bulk Import", report, "nms_unit_import_report.csv")

@bp.route("/users", methods=["GET","POST"])
@login_required
@require_roles(Role.ADMIN.value)
def users():
    """Global account creation only.

    This screen deliberately excludes unit workspace accounts and internal unit
    officers so Super Admin does not confuse View-Only Officers with Unit
    Accounts, Chief Clerks, AOs, Commanders or Signatory Officers.
    """
    if request.method == "POST":
        full_name = request.form.get("full_name","").strip()
        service_number = request.form.get("service_number","").strip().upper()
        password = request.form.get("password","")
        rank = request.form.get("rank", "").strip() or None
        specialty = request.form.get("specialty", "").strip() or None
        email = request.form.get("email", "").strip() or None
        phone = request.form.get("phone", "").strip() or None
        role = _safe_role(request.form.get("role", Role.OFFICER.value))
        unit_id = request.form.get("unit_id") or None
        clearance_level = request.form.get("clearance_level") or _rank_clearance(rank)
        if request.form.get("auto_clearance") == "on":
            clearance_level = _rank_clearance(rank)

        if role not in {Role.OFFICER.value, Role.ADMIN.value, Role.SUPER_ADMIN.value}:
            flash("Use Unit Users to create Chief Clerk, AO, Commander, Signatory or Other Officer accounts.", "warning")
            return redirect(url_for("admin.users"))
        if role == Role.SUPER_ADMIN.value and current_user.role != Role.SUPER_ADMIN.value:
            abort(403)
        if not full_name or not service_number:
            flash("Full name and service number are required", "danger")
            return redirect(url_for("admin.users"))
        if service_number.startswith("UNIT/"):
            flash("UNIT/ accounts must be created from Unit Accounts and linked to an existing unit.", "warning")
            return redirect(url_for("admin.unit_accounts"))
        generated_password = False
        if not password:
            password = secrets.token_urlsafe(10) + "Aa1!"
            generated_password = True
        ok, issues = validate_password_strength(password)
        if not ok:
            flash("Password policy failed: " + "; ".join(issues), "danger")
            return redirect(url_for("admin.users"))
        if User.query.filter_by(service_number=service_number).first():
            flash("Service number already exists", "danger")
            return redirect(url_for("admin.users"))
        if email and User.query.filter(User.email.ilike(email)).first():
            flash("Email address already exists", "danger")
            return redirect(url_for("admin.users"))

        u = User(
            full_name=full_name,
            service_number=service_number,
            rank=rank,
            specialty=specialty,
            email=email,
            phone=phone,
            appointment="View-Only Officer" if role == Role.OFFICER.value else "HQ Administrator",
            account_type="OFFICER",
            role=role,
            unit_id=unit_id,
            clearance_level=clearance_level,
            must_change_password=True,
            password_hash="tmp",
        )
        u.set_password(password)
        db.session.add(u); db.session.commit()
        log_event(current_user.id, "GLOBAL_USER_CREATED", {
            "target_module": "Global Users",
            "target_id": u.id,
            "service_number": service_number,
            "role": u.role,
            "rank": u.rank,
            "clearance": u.clearance_level,
            "unit": u.unit.code if u.unit else "",
            "account_type": u.account_type,
            "generated_password": generated_password,
            "status": "SUCCESS",
        })
        if request.form.get("email_login_details") == "on":
            if _send_login_email(u, password):
                flash("Global user created and login details sent to email", "success")
            else:
                flash("Global user created. Email was not sent because SMTP/email is not configured or delivery failed.", "warning")
        else:
            flash("Global user created" + (" with an auto-generated temporary password." if generated_password else ""), "success")
        return redirect(url_for("admin.users"))

    users = _global_user_query().order_by(User.created_at.desc()).all()
    units = Unit.query.order_by(Unit.name.asc()).all()
    role_options = [(Role.OFFICER.value, "View-Only Officer"), (Role.ADMIN.value, "HQ Admin")]
    if current_user.role == Role.SUPER_ADMIN.value:
        role_options.append((Role.SUPER_ADMIN.value, "Super Admin"))
    return render_template("admin/users.html", users=users, units=units, Role=Role, rank_options=RANK_OPTIONS, role_options=role_options)


@bp.get("/users/export.csv")
@login_required
@require_roles(Role.ADMIN.value)
def export_users_csv():
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Full Name", "Service Number", "Rank", "Role", "Account Type", "Unit", "Appointment", "Clearance", "Email", "Status", "Last Login", "Created At"])
    for u in _global_user_query().order_by(User.full_name.asc()).all():
        writer.writerow([
            u.full_name, u.service_number, u.rank or "", u.role, getattr(u, "account_type", "OFFICER"),
            u.unit.code if u.unit else "", getattr(u, "appointment", "") or "", u.clearance_level, getattr(u, "email", "") or "",
            "ACTIVE" if u.is_active_flag else "DISABLED",
            u.last_login_at.isoformat(sep=" ") if u.last_login_at else "",
            u.created_at.isoformat(sep=" ") if u.created_at else "",
        ])
    log_event(current_user.id, "USERS_EXPORTED_CSV", "User directory exported")
    return Response(output.getvalue(), mimetype="text/csv", headers={"Content-Disposition": "attachment; filename=nms_global_users_export.csv"})


@bp.route("/users/<int:user_id>/edit", methods=["GET","POST"])
@login_required
@require_roles(Role.ADMIN.value)
def edit_user(user_id):
    u = User.query.get_or_404(user_id)
    if u.role == Role.SUPER_ADMIN.value and current_user.role != Role.SUPER_ADMIN.value:
        abort(403)

    if request.method == "POST":
        u.full_name = request.form.get("full_name","").strip() or u.full_name
        u.service_number = request.form.get("service_number","").strip() or u.service_number
        role = _safe_role(request.form.get("role", u.role))
        u.rank = request.form.get("rank", "").strip() or None
        u.specialty = request.form.get("specialty", "").strip() or None
        u.email = request.form.get("email", "").strip() or None
        u.phone = request.form.get("phone", "").strip() or None
        # Global user edit screen must not turn accounts into Unit Accounts or unit workflow officers.
        requested_appointment = request.form.get("appointment", "").strip() or None
        if normalize_unit_appointment(requested_appointment) in {"CHIEF_CLERK", "ADMIN_OFFICER", "SIGNATORY_OFFICER", "COMMANDER", "OTHER_OFFICER"}:
            flash("Use Unit Users to manage Chief Clerk, AO, Commander, Signatory or Other Officer appointments.", "warning")
            return redirect(url_for("admin.edit_user", user_id=u.id))
        u.appointment = requested_appointment
        u.account_type = "OFFICER"
        if role not in {Role.OFFICER.value, Role.ADMIN.value, Role.SUPER_ADMIN.value}:
            flash("Use Unit Users for unit workflow roles.", "warning")
            return redirect(url_for("admin.edit_user", user_id=u.id))
        u.role = role
        u.unit_id = request.form.get("unit_id") or None
        u.clearance_level = request.form.get("clearance_level") or u.clearance_level
        if request.form.get("auto_clearance") == "on":
            u.clearance_level = _rank_clearance(u.rank, u.clearance_level)
        active = request.form.get("is_active_flag")
        u.is_active_flag = True if active == "on" else False
        db.session.commit()
        log_event(current_user.id, "USER_UPDATED", u.service_number)
        flash("User updated", "success")
        return redirect(url_for("admin.users"))

    units = Unit.query.order_by(Unit.name.asc()).all()
    role_options = [(Role.OFFICER.value, "View-Only Officer"), (Role.ADMIN.value, "HQ Admin")]
    if current_user.role == Role.SUPER_ADMIN.value:
        role_options.append((Role.SUPER_ADMIN.value, "Super Admin"))
    return render_template("admin/user_edit.html", u=u, units=units, Role=Role, rank_options=RANK_OPTIONS, role_options=role_options)

@bp.post("/users/<int:user_id>/reset_password")
@login_required
@require_roles(Role.ADMIN.value)
def reset_user_password(user_id):
    u = User.query.get_or_404(user_id)
    if u.role == Role.SUPER_ADMIN.value and current_user.role != Role.SUPER_ADMIN.value:
        abort(403)
    password = request.form.get("password","")
    if not password:
        flash("Password required", "danger")
        return redirect(url_for("admin.edit_user", user_id=user_id))
    ok, issues = validate_password_strength(password)
    if not ok:
        flash("Password policy failed: " + "; ".join(issues), "danger")
        return redirect(url_for("admin.edit_user", user_id=user_id))
    u.set_password(password)
    db.session.commit()
    log_event(current_user.id, "USER_PASSWORD_RESET", u.service_number)
    flash("Password reset", "success")
    return redirect(url_for("admin.edit_user", user_id=user_id))

@bp.route("/units/<int:unit_id>/edit", methods=["GET","POST"])
@login_required
@require_roles(Role.ADMIN.value)
def edit_unit(unit_id):
    unit = Unit.query.get_or_404(unit_id)
    if request.method == "POST":
        name = request.form.get("name","").strip()
        code = request.form.get("code","").strip().upper()
        level = request.form.get("level", unit.level)
        parent_id = request.form.get("parent_id") or None

        if not name or not code:
            flash("Name and code required", "danger")
            return redirect(url_for("admin.edit_unit", unit_id=unit_id))

        # prevent parent loop
        if parent_id and int(parent_id) == unit.id:
            flash("Unit cannot be parent of itself", "danger")
            return redirect(url_for("admin.edit_unit", unit_id=unit_id))

        unit.name = name
        unit.code = code
        unit.level = level
        unit.parent_id = int(parent_id) if parent_id else None

        db.session.commit()
        log_event(current_user.id, "UNIT_UPDATED", f"{unit.code}:{unit.name}")
        flash("Unit updated", "success")
        return redirect(url_for("admin.units"))

    units = Unit.query.order_by(Unit.name.asc()).all()
    return render_template("admin/unit_edit.html", unit=unit, units=units, UnitLevel=UnitLevel)

@bp.post("/units/<int:unit_id>/delete")
@login_required
@require_roles(Role.ADMIN.value)
def delete_unit(unit_id):
    unit = Unit.query.get_or_404(unit_id)
    if User.query.filter_by(unit_id=unit.id).first():
        flash("Cannot delete unit: users are assigned to it.", "danger")
        return redirect(url_for("admin.units"))
    if unit.children.count() > 0:
        flash("Cannot delete unit: it has child units.", "danger")
        return redirect(url_for("admin.units"))

    db.session.delete(unit)
    db.session.commit()
    log_event(current_user.id, "UNIT_DELETED", f"{unit.code}:{unit.name}")
    flash("Unit deleted", "success")
    return redirect(url_for("admin.units"))


@bp.route("/unit-accounts", methods=["GET", "POST"])
@login_required
@require_roles(Role.ADMIN.value)
def unit_accounts():
    """Create and manage unit workspace logins linked to existing units only."""
    if request.method == "POST":
        unit_id = request.form.get("unit_id", type=int)
        unit = Unit.query.get(unit_id) if unit_id else None

        if not unit:
            flash("Select an existing unit before creating a unit account.", "danger")
            return redirect(url_for("admin.unit_accounts"))

        email = (request.form.get("unit_email") or "").strip().lower() or None
        password = (request.form.get("unit_password") or "").strip() or None

        acct, result = _create_unit_workspace_account(unit, email, password)

        if not acct:
            flash(result, "warning")
            return redirect(url_for("admin.unit_accounts"))

        flash(
            f"Unit workspace account created for {unit.code}. "
            f"Login: {acct.service_number}. Temporary password: {result}",
            "success"
        )
        return redirect(url_for("admin.unit_officers", unit_id=unit.id))

    q = (request.args.get("q") or "").strip()

    units_query = Unit.query

    if q:
        like = f"%{q}%"
        units_query = units_query.filter(
            (Unit.name.ilike(like)) |
            (Unit.code.ilike(like)) |
            (Unit.level.ilike(like))
        )

    units = units_query.order_by(Unit.name.asc()).all()

    # Optimized: fetch all unit accounts in one query instead of querying once per unit.
    unit_ids = [u.id for u in units]

    accounts = {}

    if unit_ids:
        unit_accounts = (
            User.query
            .filter(
                User.account_type == "UNIT",
                User.unit_id.in_(unit_ids)
            )
            .order_by(User.created_at.asc())
            .all()
        )

        for account in unit_accounts:
            # Keep the first/oldest account per unit, matching the old logic.
            if account.unit_id not in accounts:
                accounts[account.unit_id] = account

    return render_template(
        "admin/unit_accounts.html",
        units=units,
        accounts=accounts,
        search=q
    )

@bp.post("/unit-accounts/<int:user_id>/toggle")
@login_required
@require_roles(Role.ADMIN.value)
def toggle_unit_account(user_id):
    acct = User.query.get_or_404(user_id)
    if acct.account_type != "UNIT":
        abort(400)
    acct.is_active_flag = not bool(acct.is_active_flag)
    db.session.commit()
    log_event(current_user.id, "UNIT_WORKSPACE_ACCOUNT_STATUS_CHANGED", {"account_id": acct.id, "unit_id": acct.unit_id, "active": acct.is_active_flag})
    flash(f"Unit account {acct.service_number} is now {'active' if acct.is_active_flag else 'disabled'}.", "success")
    return redirect(url_for("admin.unit_accounts"))

@bp.post("/unit-accounts/<int:user_id>/reset-password")
@login_required
@require_roles(Role.ADMIN.value)
def reset_unit_account_password(user_id):
    acct = User.query.get_or_404(user_id)
    if acct.account_type != "UNIT":
        abort(400)
    password = _suggest_export_password("UNIT")
    acct.set_password(password)
    acct.must_change_password = True
    db.session.commit()
    sent = _send_login_email(acct, password)
    log_event(current_user.id, "UNIT_WORKSPACE_ACCOUNT_PASSWORD_RESET", {"account_id": acct.id, "unit_id": acct.unit_id, "email_sent": sent})
    flash(("Unit account password reset and emailed." if sent else f"Unit account password reset. Temporary password: {password}"), "success")
    return redirect(url_for("admin.unit_accounts"))

@bp.route("/units", methods=["GET","POST"])
@login_required
@require_roles(Role.ADMIN.value)
def units():
    if request.method == "POST":
        name = request.form.get("name","").strip()
        code = request.form.get("code","").strip().upper()
        if not name or not code:
            flash("Name and code required", "danger")
            return redirect(url_for("admin.units"))
        if Unit.query.filter((Unit.name==name) | (Unit.code==code)).first():
            flash("Unit name/code already exists", "danger")
            return redirect(url_for("admin.units"))
        level = request.form.get("level", UnitLevel.UNIT.value)
        parent_id = request.form.get("parent_id") or None
        u = Unit(name=name, code=code, level=level, parent_id=int(parent_id) if parent_id else None)
        db.session.add(u); db.session.commit()

        log_event(current_user.id, "UNIT_CREATED", f"{code}:{name}")
        flash("Unit created. Create its workspace login from Unit Accounts when ready.", "success")
        return redirect(url_for("admin.unit_accounts"))

    search = (request.args.get("q") or "").strip()
    page = request.args.get("page", default=1, type=int)
    per_page = 12

    units_query = Unit.query
    if search:
        like = f"%{search}%"
        units_query = units_query.filter(
            (Unit.name.ilike(like)) |
            (Unit.code.ilike(like)) |
            (Unit.level.ilike(like))
        )

    pagination = units_query.order_by(Unit.name.asc()).paginate(page=page, per_page=per_page, error_out=False)
    all_units = Unit.query.order_by(Unit.name.asc()).all()
    return render_template(
        "admin/units.html",
        units=pagination.items,
        pagination=pagination,
        total_units=units_query.count(),
        all_units=all_units,
        search=search,
        UnitLevel=UnitLevel,
    )

@bp.post("/units/<int:unit_id>/create-account")
@login_required
@require_roles(Role.ADMIN.value)
def create_unit_account(unit_id):
    """Backward-compatible quick action. Unit accounts still link to an existing unit."""
    unit = Unit.query.get_or_404(unit_id)
    acct, result = _create_unit_workspace_account(unit, request.form.get("unit_email"), request.form.get("unit_password"))
    if not acct:
        flash(result, "warning")
        return redirect(url_for("admin.unit_accounts"))
    flash(f"Unit workspace account created for {unit.code}. Login: {acct.service_number}. Temporary password: {result}", "success")
    return redirect(url_for("admin.unit_officers", unit_id=unit.id))

@bp.route("/channels", methods=["GET","POST"])
@login_required
@require_roles(Role.ADMIN.value)
def channels():
    if request.method == "POST":
        name = request.form.get("name","").strip()
        scope = request.form.get("scope","UNIT")
        unit_id = request.form.get("unit_id") or None
        unit_ids = [int(x) for x in request.form.getlist("unit_ids") if str(x).isdigit()]
        officer_ids = [int(x) for x in request.form.getlist("officer_ids") if str(x).isdigit()]
        clearance_level = request.form.get("clearance_level") or "RESTRICTED"
        description = (request.form.get("description") or "").strip()
        if not name:
            flash("Channel name required", "danger")
            return redirect(url_for("admin.channels"))
        c = Channel(name=name, scope=scope, unit_id=unit_id, created_by_id=current_user.id, classification_level=clearance_level, description=description)
        c.units = Unit.query.filter(Unit.id.in_(unit_ids)).all() if unit_ids else []
        c.members = User.query.filter(User.id.in_(officer_ids)).all() if officer_ids else []
        db.session.add(c); db.session.commit()
        log_event(current_user.id, "CHANNEL_CREATED", f"{name} ({scope}) units={unit_ids} officers={officer_ids} class={clearance_level}")
        flash("Channel created", "success")
        return redirect(url_for("admin.channels"))

    channels = Channel.query.order_by(Channel.created_at.desc()).all()
    units = Unit.query.order_by(Unit.name.asc()).all()
    users = User.query.filter_by(is_active_flag=True).order_by(User.full_name.asc()).all()
    return render_template("admin/channels.html", channels=channels, units=units, users=users)


@bp.get("/broadcasts")
@login_required
@require_roles(Role.ADMIN.value)
def broadcasts_admin():
    """Admin view of broadcasts with delivery/read ticks + print links."""
    bcasts = Broadcast.query.order_by(Broadcast.created_at.desc()).limit(300).all()
    stats_by_id = {}
    for b in bcasts:
        receipts = BroadcastReceipt.query.filter_by(broadcast_id=b.id).all()
        not_seen = sum(1 for r in receipts if not r.received_at)
        received = sum(1 for r in receipts if r.received_at and not r.read_at)
        read = sum(1 for r in receipts if r.read_at)
        acked = 0
        if b.requires_ack:
            acked = BroadcastAck.query.filter(BroadcastAck.broadcast_id == b.id, BroadcastAck.acked_at.isnot(None)).count()
        stats_by_id[b.id] = {"not_seen": not_seen, "received": received, "read": read, "acked": acked}
    return render_template("admin/broadcasts.html", broadcasts=bcasts, stats_by_id=stats_by_id)


@bp.route("/channels/<int:channel_id>/edit", methods=["GET","POST"])
@login_required
@require_roles(Role.ADMIN.value)
def edit_channel(channel_id):
    c = Channel.query.get_or_404(channel_id)
    if request.method == "POST":
        c.name = (request.form.get("name") or c.name).strip()
        c.scope = request.form.get("scope", c.scope)
        unit_id = request.form.get("unit_id") or None
        unit_ids = [int(x) for x in request.form.getlist("unit_ids") if str(x).isdigit()]
        officer_ids = [int(x) for x in request.form.getlist("officer_ids") if str(x).isdigit()]
        clearance_level = request.form.get("clearance_level") or "RESTRICTED"
        description = (request.form.get("description") or "").strip()
        c.unit_id = int(unit_id) if unit_id else None
        db.session.commit()
        log_event(current_user.id, "CHANNEL_UPDATED", f"{c.id}:{c.name}")
        flash("Channel updated", "success")
        return redirect(url_for("admin.channels"))
    units = Unit.query.order_by(Unit.name.asc()).all()
    users = User.query.filter_by(is_active_flag=True).order_by(User.full_name.asc()).all()
    return render_template("admin/channel_edit.html", channel=c, units=units, users=users)


@bp.post("/channels/<int:channel_id>/delete")
@login_required
@require_roles(Role.ADMIN.value)
def delete_channel(channel_id):
    c = Channel.query.get_or_404(channel_id)
    db.session.delete(c)
    db.session.commit()
    log_event(current_user.id, "CHANNEL_DELETED", f"{c.id}:{c.name}")
    flash("Channel deleted", "success")
    return redirect(url_for("admin.channels"))


@bp.get("/units/tree")
@login_required
@require_roles(Role.ADMIN.value)
def units_tree():
    # Simple tree view for hierarchy sanity checks
    units = Unit.query.order_by(Unit.name.asc()).all()
    children = {}
    by_id = {u.id: u for u in units}
    roots = []
    for u in units:
        pid = u.parent_id
        if pid and pid in by_id:
            children.setdefault(pid, []).append(u)
        else:
            roots.append(u)
    # sort children lists
    for pid in list(children.keys()):
        children[pid] = sorted(children[pid], key=lambda x: (x.level or "", x.name or ""))
    roots = sorted(roots, key=lambda x: (x.level or "", x.name or ""))
    return render_template("admin/units_tree.html", roots=roots, children=children)



def _audit_query_from_request():
    q = AuditEvent.query.outerjoin(User, AuditEvent.actor_id == User.id)
    action = (request.args.get("action") or "").strip()
    actor = (request.args.get("actor") or "").strip()
    date_from = (request.args.get("date_from") or "").strip()
    date_to = (request.args.get("date_to") or "").strip()
    text = (request.args.get("q") or "").strip()
    unit_id = request.args.get("unit_id", type=int)
    if action:
        q = q.filter(AuditEvent.action.ilike(f"%{action}%"))
    if text:
        like = f"%{text}%"
        q = q.filter((AuditEvent.action.ilike(like)) | (AuditEvent.details.ilike(like)) | (AuditEvent.event_hash.ilike(like)))
    if actor:
        q = q.filter(
            (User.full_name.ilike(f"%{actor}%")) | (User.service_number.ilike(f"%{actor}%")) | (User.rank.ilike(f"%{actor}%"))
        )
    if unit_id:
        unit = Unit.query.get(unit_id)
        if unit:
            unit_id_text = f'"unit_id": {unit.id}'
            compact_unit_id_text = f'"unit_id":{unit.id}'
            q = q.filter(
                (User.unit_id == unit.id)
                | (AuditEvent.details.ilike(f"%{unit_id_text}%"))
                | (AuditEvent.details.ilike(f"%{compact_unit_id_text}%"))
                | (AuditEvent.details.ilike(f"%{unit.code}%"))
                | (AuditEvent.details.ilike(f"%{unit.name}%"))
            )
    try:
        if date_from:
            q = q.filter(AuditEvent.created_at >= datetime.strptime(date_from, "%Y-%m-%d"))
        if date_to:
            end = datetime.strptime(date_to, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
            q = q.filter(AuditEvent.created_at <= end)
    except ValueError:
        pass
    return q

def _audit_details_map(a):
    """Return audit details as a safe dictionary for templates/CSV.

Older audit rows may contain plain text, numbers, booleans, or JSON lists
inside AuditEvent.details. Jinja expects a mapping with .get(), so this
normalizes every legacy/new value into a dict and prevents admin logs from
crashing.
"""
    raw = a.details
    if raw is None or raw == "":
        return {}
    try:
        parsed = json.loads(raw) if isinstance(raw, str) else raw
    except Exception:
        return {"summary": str(raw)}
    if isinstance(parsed, dict):
        return parsed
    if isinstance(parsed, list):
        return {"summary": ", ".join(str(x) for x in parsed)}
    return {"summary": str(parsed)}

def _audit_category(action):
    action = (action or "").upper()
    if "LOGIN" in action or "LOGOUT" in action or "SESSION" in action or "LOCK" in action:
        return "Access & Sessions"
    if "PRINT" in action or "EXPORT" in action or "DOWNLOAD" in action:
        return "Print & Export"
    if "SIGNAL" in action or "BROADCAST" in action or "ACK" in action:
        return "Signals"
    if "USER" in action or "PASSWORD" in action or "ROLE" in action or "UNIT" in action:
        return "Users & Units"
    if "SETTING" in action or "MAINTENANCE" in action or "BACKUP" in action or "RETENTION" in action:
        return "System"
    return "Other"

def _audit_status(a, d):
    status = (d.get("status") or "SUCCESS").upper() if isinstance(d, dict) else "SUCCESS"
    if "FAIL" in (a.action or "").upper() or "DENIED" in (a.action or "").upper() or "BLOCK" in (a.action or "").upper():
        status = status if status != "SUCCESS" else "FAILED"
    return status

def _audit_display_rows(audits):
    today = datetime.utcnow().date()
    groups = {"Today": [], "Yesterday": [], "This Week": [], "Older": []}
    rows = []
    for a in audits:
        d = _audit_details_map(a)
        if not isinstance(d, dict):
            d = {"summary": str(d)}
        actor = a.actor
        created = a.created_at or datetime.utcnow()
        row = {
            "id": a.id,
            "created_at": created,
            "created_label": created.strftime("%Y-%m-%d %H:%M:%S"),
            "action": a.action,
            "category": _audit_category(a.action),
            "status": _audit_status(a, d),
            "actor_name": d.get("actor_name") or (actor.full_name if actor else "System"),
            "rank": d.get("rank") or (actor.rank if actor else ""),
            "service_number": d.get("service_number") or (actor.service_number if actor else ""),
            "role": d.get("role") or (actor.role if actor else ""),
            "unit": d.get("unit") or (actor.unit.name if actor and actor.unit else ""),
            "unit_id": d.get("unit_id") or d.get("sender_unit_id") or d.get("recipient_unit_id") or (actor.unit_id if actor else None),
            "unit_code": d.get("unit_code") or (actor.unit.code if actor and actor.unit else ""),
            "ip": d.get("ip", ""),
            "device": d.get("device", ""),
            "target_module": d.get("target_module", ""),
            "target_id": d.get("target_id", ""),
            "signal_ref": d.get("signal_ref", ""),
            "classification": d.get("classification", ""),
            "priority": d.get("priority", ""),
            "print_export_id": d.get("print_id") or d.get("export_id") or "",
            "pages": d.get("pages", ""),
            "remarks": d.get("remarks") or d.get("summary") or "",
            "event_hash": a.event_hash or "",
            "prev_hash": a.prev_hash or "ROOT",
        }
        event_date = created.date()
        if event_date == today:
            groups["Today"].append(row)
        elif event_date == today - timedelta(days=1):
            groups["Yesterday"].append(row)
        elif event_date >= today - timedelta(days=7):
            groups["This Week"].append(row)
        else:
            groups["Older"].append(row)
        rows.append(row)
    return rows, groups



def _audit_folder_name(row):
    action = (row.get("action") or "").upper()
    if "WORKFLOW" in action or "ACK" in action or "SIGNAL" in action or "BROADCAST" in action:
        return "Signal & Workflow Actions"
    if "DIRECT_MESSAGE" in action or "MESSAGE" in action or "CHANNEL" in action:
        return "Messages"
    if "USER" in action or "PASSWORD" in action or "ROLE" in action or "UNIT" in action or "OFFICER" in action:
        return "Users & Unit Management"
    if "LOGIN" in action or "SESSION" in action or "LOCK" in action or "DENIED" in action or "FAIL" in action:
        return "Security & Access"
    if "SETTING" in action or "BACKUP" in action or "EXPORT" in action or "DOWNLOAD" in action or "PRINT" in action:
        return "System, Print & Export"
    return "Other Events"

def _audit_folder_groups(rows):
    folders = OrderedDict()
    for name in ["Signal & Workflow Actions", "Messages", "Users & Unit Management", "Security & Access", "System, Print & Export", "Other Events"]:
        folders[name] = []
    for row in rows:
        folders.setdefault(_audit_folder_name(row), []).append(row)
    return OrderedDict((k, v) for k, v in folders.items() if v)

def _audit_unit_groups(rows):
    groups = OrderedDict()
    for row in rows:
        unit_name = row.get("unit") or "System / No Unit"
        unit_key = row.get("unit_id") or unit_name
        if unit_key not in groups:
            groups[unit_key] = {"name": unit_name, "code": row.get("unit_code") or "", "rows": [], "folders": OrderedDict()}
        groups[unit_key]["rows"].append(row)
    for data in groups.values():
        data["folders"] = _audit_folder_groups(data["rows"])
    return groups


def _unit_audit_scope_or_403(unit_id=None):
    """Return the unit ID the current user is allowed to audit.

    HQ admins may inspect any unit. Unit AO/Commander/unit-management accounts
    are restricted to their own unit only.
    """
    return _unit_scope_or_403(unit_id)


def _unit_audit_query(unit_id: int):
    """Audit rows connected to a unit.

    Primary match is actor.unit_id. A secondary text match catches system/HQ
    actions that wrote the unit_id, unit code or unit name into the audit
    payload even when the actor was outside the unit.
    """
    unit = Unit.query.get_or_404(unit_id)
    q = AuditEvent.query.outerjoin(User, AuditEvent.actor_id == User.id)
    unit_id_text = f'"unit_id": {unit.id}'
    compact_unit_id_text = f'"unit_id":{unit.id}'
    return q.filter(
        (User.unit_id == unit.id)
        | (AuditEvent.details.ilike(f"%{unit_id_text}%"))
        | (AuditEvent.details.ilike(f"%{compact_unit_id_text}%"))
        | (AuditEvent.details.ilike(f"%{unit.code}%"))
        | (AuditEvent.details.ilike(f"%{unit.name}%"))
    )


@bp.get("/unit-audit")
@login_required
def unit_audit():
    selectable_units = Unit.query.order_by(Unit.name.asc()).all() if current_user.role in (Role.ADMIN.value, Role.SUPER_ADMIN.value) else []
    requested_unit_id = request.args.get("unit_id", type=int) if current_user.role in (Role.ADMIN.value, Role.SUPER_ADMIN.value) else current_user.unit_id
    if not requested_unit_id and selectable_units:
        requested_unit_id = selectable_units[0].id
    unit_id = _unit_audit_scope_or_403(requested_unit_id)
    unit = Unit.query.get_or_404(unit_id)

    base_q = _unit_audit_query(unit.id)
    action = (request.args.get("action") or "").strip()
    text = (request.args.get("q") or "").strip()
    date_from = (request.args.get("date_from") or "").strip()
    date_to = (request.args.get("date_to") or "").strip()
    if action:
        base_q = base_q.filter(AuditEvent.action.ilike(f"%{action}%"))
    if text:
        like = f"%{text}%"
        base_q = base_q.filter((AuditEvent.action.ilike(like)) | (AuditEvent.details.ilike(like)) | (AuditEvent.event_hash.ilike(like)) | (AuditEvent.payload_sha256.ilike(like)))
    try:
        if date_from:
            base_q = base_q.filter(AuditEvent.created_at >= datetime.strptime(date_from, "%Y-%m-%d"))
        if date_to:
            end = datetime.strptime(date_to, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
            base_q = base_q.filter(AuditEvent.created_at <= end)
    except ValueError:
        pass

    audits = base_q.order_by(AuditEvent.created_at.desc()).limit(500).all()
    rows, grouped_audits = _audit_display_rows(audits)
    audit_folders = _audit_folder_groups(rows)
    summary = {
        "total": len(rows),
        "workflow": sum(1 for r in rows if "WORKFLOW" in (r.get("action") or "")),
        "users": sum(1 for r in rows if r.get("category") == "Users & Units"),
        "messages": sum(1 for r in rows if "DIRECT_MESSAGE" in (r.get("action") or "") or "MESSAGE" in (r.get("action") or "")),
        "failed": sum(1 for r in rows if r.get("status") != "SUCCESS"),
    }
    action_options = sorted({r["action"] for r in rows if r.get("action")})[:80]
    return render_template(
        "admin/unit_audit.html",
        unit=unit,
        units=selectable_units,
        audit_rows=rows,
        grouped_audits=grouped_audits,
        audit_folders=audit_folders,
        summary=summary,
        action_options=action_options,
        request_args=request.args,
    )


@bp.get("/unit-audit/export.csv")
@login_required
def export_unit_audit_csv():
    selectable_allowed = current_user.role in (Role.ADMIN.value, Role.SUPER_ADMIN.value)
    requested_unit_id = request.args.get("unit_id", type=int) if selectable_allowed else current_user.unit_id
    unit_id = _unit_audit_scope_or_403(requested_unit_id)
    unit = Unit.query.get_or_404(unit_id)
    q = _unit_audit_query(unit.id)
    action = (request.args.get("action") or "").strip()
    text = (request.args.get("q") or "").strip()
    date_from = (request.args.get("date_from") or "").strip()
    date_to = (request.args.get("date_to") or "").strip()
    if action:
        q = q.filter(AuditEvent.action.ilike(f"%{action}%"))
    if text:
        like = f"%{text}%"
        q = q.filter((AuditEvent.action.ilike(like)) | (AuditEvent.details.ilike(like)) | (AuditEvent.event_hash.ilike(like)) | (AuditEvent.payload_sha256.ilike(like)))
    try:
        if date_from:
            q = q.filter(AuditEvent.created_at >= datetime.strptime(date_from, "%Y-%m-%d"))
        if date_to:
            end = datetime.strptime(date_to, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
            q = q.filter(AuditEvent.created_at <= end)
    except ValueError:
        pass
    audits = q.order_by(AuditEvent.created_at.desc()).limit(5000).all()
    rows, _groups = _audit_display_rows(audits)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Date/Time UTC", "Action", "Status", "Officer", "Rank", "Service Number", "Unit", "Signal Ref", "Module", "Remarks", "Event Hash", "Previous Hash"])
    for r in rows:
        writer.writerow([r.get("created_label"), r.get("action"), r.get("status"), r.get("actor_name"), r.get("rank"), r.get("service_number"), r.get("unit"), r.get("signal_ref"), r.get("target_module"), r.get("remarks"), r.get("event_hash"), r.get("prev_hash")])
    log_event(current_user.id, "UNIT_AUDIT_EXPORTED_CSV", {"unit_id": unit.id, "unit": unit.code, "records": len(rows), "status": "SUCCESS"})
    filename = f"unit_audit_{unit.code or unit.id}.csv".replace(" ", "_")
    return Response(output.getvalue().encode("utf-8-sig"), mimetype="text/csv", headers={"Content-Disposition": f"attachment; filename={filename}"})

@bp.get("/logs")
@login_required
@require_roles(Role.ADMIN.value)
def logs():
    audits = _audit_query_from_request().order_by(AuditEvent.created_at.desc()).limit(1500).all()
    rows, grouped_audits = _audit_display_rows(audits)
    audit_unit_groups = _audit_unit_groups(rows)
    unit_options = Unit.query.order_by(Unit.name.asc()).all()
    summary = {
        "total": len(rows),
        "access": sum(1 for r in rows if r["category"] == "Access & Sessions"),
        "signals": sum(1 for r in rows if r["category"] == "Signals"),
        "print_export": sum(1 for r in rows if r["category"] == "Print & Export"),
        "failed": sum(1 for r in rows if r["status"] != "SUCCESS"),
        "units": len([k for k in audit_unit_groups.keys() if k != "System / No Unit"]),
        "users": len({r.get("service_number") for r in rows if r.get("service_number")}),
    }
    action_options = sorted({r["action"] for r in rows if r.get("action")})[:80]
    return render_template(
        "admin/logs.html",
        audits=audits,
        audit_rows=rows,
        grouped_audits=grouped_audits,
        audit_unit_groups=audit_unit_groups,
        audit_summary=summary,
        unit_options=unit_options,
        action_options=action_options,
        details_map=_audit_details_map,
    )

@bp.get("/logs/export.csv")
@login_required
@require_roles(Role.ADMIN.value)
def export_audit_logs_csv():
    audits = _audit_query_from_request().order_by(AuditEvent.created_at.desc()).limit(10000).all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Date/Time UTC", "Action", "Status", "Officer Name", "Rank", "Service Number", "Role", "Unit",
        "IP Address", "Device/Browser", "Target Module", "Target ID", "Signal Reference",
        "Classification", "Priority", "Print/Export ID", "Pages", "Remarks", "Event Hash", "Previous Hash"
    ])
    for a in audits:
        d = _audit_details_map(a)
        actor = a.actor
        writer.writerow([
            a.created_at.strftime("%Y-%m-%d %H:%M:%S") if a.created_at else "",
            a.action,
            d.get("status", "SUCCESS"),
            d.get("actor_name") or (actor.full_name if actor else "System"),
            d.get("rank") or (actor.rank if actor else ""),
            d.get("service_number") or (actor.service_number if actor else ""),
            d.get("role") or (actor.role if actor else ""),
            d.get("unit") or (actor.unit.name if actor and actor.unit else ""),
            d.get("ip", ""),
            d.get("device", ""),
            d.get("target_module", ""),
            d.get("target_id", ""),
            d.get("signal_ref", ""),
            d.get("classification", ""),
            d.get("priority", ""),
            d.get("print_id") or d.get("export_id", ""),
            d.get("pages", ""),
            d.get("remarks") or d.get("summary", ""),
            a.event_hash or "",
            a.prev_hash or "ROOT",
        ])
    log_event(current_user.id, "AUDIT_LOG_EXPORTED_CSV", {"target_module":"Audit Logs", "records": len(audits), "status":"SUCCESS"})
    csv_bytes = output.getvalue().encode("utf-8-sig")
    return Response(csv_bytes, mimetype="text/csv", headers={"Content-Disposition":"attachment; filename=audit_logs.csv"})


@bp.get("/system_logs")
@login_required
@require_roles(Role.SUPER_ADMIN.value)
def system_logs():
    """View the last lines of instance/logs/app.log (server-side errors/info)."""
    import os
    from flask import current_app

    log_path = os.path.join(current_app.instance_path, "logs", "app.log")
    lines = []
    if os.path.exists(log_path):
        try:
            with open(log_path, "r", encoding="utf-8", errors="replace") as f:
                lines = f.readlines()[-400:]
        except Exception:
            lines = ["Unable to read log file. Check file permissions."]
    else:
        lines = ["Log file not found yet. It will appear after the server writes its first log entry."]

    return render_template("admin/system_logs.html", lines=lines)

@bp.route("/settings", methods=["GET","POST"])
@login_required
@require_roles(Role.ADMIN.value)
def settings():
    # General Admin / HQ and Super Admin can manage platform lock/settings.
    s = Settings.get()
    if request.method == "POST":
        s.platform_name = (request.form.get("platform_name", "").strip() or "NAF Secure Messaging")
        before_lock = bool(s.maintenance_mode)
        s.maintenance_mode = True if request.form.get("maintenance_mode") == "on" else False
        s.maintenance_banner = (request.form.get("maintenance_banner","").strip() or "")
        s.maintenance_reason = (request.form.get("maintenance_reason","").strip() or "Scheduled maintenance")
        s.maintenance_expected_return = (request.form.get("maintenance_expected_return","").strip() or "")
        s.dark_mode_default = True if request.form.get("dark_mode_default") == "on" else False
        s.allowed_extensions = (request.form.get("allowed_extensions","").strip() or s.allowed_extensions)
        s.storage_backend = (request.form.get("storage_backend","local").strip() or "local")
        s.retention_days = int(request.form.get("retention_days") or 90)
        s.allow_direct_messages = True if request.form.get("allow_direct_messages") == "on" else False
        s.allow_attachments = True if request.form.get("allow_attachments") == "on" else False
        s.max_attachment_mb = int(request.form.get("max_attachment_mb") or 10)
        s.session_timeout_minutes = int(request.form.get("session_timeout_minutes") or 60)
        s.password_min_length = int(request.form.get("password_min_length") or 12)
        s.require_password_complexity = True if request.form.get("require_password_complexity") == "on" else False
        s.failed_login_limit = int(request.form.get("failed_login_limit") or 5)
        s.lockout_minutes = int(request.form.get("lockout_minutes") or 15)
        s.audit_retention_days = int(request.form.get("audit_retention_days") or 365)
        s.allow_signal_print = True if request.form.get("allow_signal_print") == "on" else False
        s.allow_signal_download = True if request.form.get("allow_signal_download") == "on" else False
        s.broadcast_escalation_minutes = int(request.form.get("broadcast_escalation_minutes") or 30)
        db.session.commit()
        if before_lock != bool(s.maintenance_mode):
            log_event(current_user.id, "SYSTEM_LOCK_ENABLED" if s.maintenance_mode else "SYSTEM_LOCK_DISABLED", s.maintenance_banner or s.maintenance_reason)
        log_event(current_user.id, "SETTINGS_UPDATED", "messaging")
        flash("Settings saved", "success")
        return redirect(url_for("admin.settings"))

    return render_template("admin/settings.html", settings=s)


@bp.post("/audit/prune")
@login_required
@require_roles(Role.ADMIN.value)
def prune_audit_logs():
    from datetime import timedelta
    s = Settings.get()
    days = max(30, int(getattr(s, "audit_retention_days", 365) or 365))
    cutoff = datetime.utcnow() - timedelta(days=days)
    deleted = AuditEvent.query.filter(AuditEvent.created_at < cutoff).delete(synchronize_session=False)
    db.session.commit()
    log_event(current_user.id, "AUDIT_RETENTION_PRUNED", f"Deleted {deleted} audit events older than {days} days")
    flash(f"Audit retention cleanup completed. {deleted} old record(s) removed.", "success")
    return redirect(url_for("admin.settings"))


@bp.get("/sessions")
@login_required
@require_roles(Role.ADMIN.value)
def sessions():
    # Admin can view sessions (SUPER_ADMIN can revoke any; ADMIN can revoke within their unit subtree)
    q = UserSession.query.order_by(UserSession.created_at.desc())
    sessions = q.limit(200).all()
    users = {u.id: u for u in User.query.filter(User.id.in_([s.user_id for s in sessions])).all()} if sessions else {}
    return render_template("admin/sessions.html", sessions=sessions, users=users)

@bp.post("/sessions/<int:sid>/revoke")
@login_required
@require_roles(Role.ADMIN.value)
def sessions_revoke(sid):
    s = UserSession.query.get_or_404(sid)
    # Restrict: only SUPER_ADMIN can revoke across the org
    if current_user.role != Role.SUPER_ADMIN.value and s.user_id != current_user.id:
        # allow admins to revoke sessions in same unit subtree
        target = User.query.get(s.user_id)
        if not target or (current_user.unit_id and target.unit_id and not (target.unit_id==current_user.unit_id)):
            abort(403)
    from datetime import datetime, timedelta
    s.revoked_at = datetime.utcnow()
    db.session.commit()
    log_event(current_user.id, "SESSION_REVOKED", f"user_id={s.user_id}")
    flash("Session revoked.", "success")
    return redirect(url_for("admin.sessions"))

@bp.route("/backup", methods=["GET", "POST"])
@login_required
@require_roles(Role.SUPER_ADMIN.value)
def backup():
    export_id = f"NMS-DB-RECOVERY-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}"
    if request.method == "GET":
        return render_template(
            "protected_export_password.html",
            export_title="Database Recovery File",
            export_id=export_id,
            suggested_password=_suggest_export_password("DBREC"),
            action_url=url_for("admin.backup"),
            cancel_url=url_for("admin.settings"),
            hidden_fields={"export_id": export_id},
        )
    password, err = _resolve_export_password()
    if err:
        flash(err, "warning")
        return redirect(url_for("admin.backup"))
    uri = current_app.config.get("SQLALCHEMY_DATABASE_URI", "")
    if uri.startswith("sqlite:///"):
        db_path = uri.replace("sqlite:///", "")
        if not os.path.isabs(db_path):
            db_path = os.path.join(current_app.instance_path, db_path)
        if os.path.exists(db_path):
            with open(db_path, "rb") as fh:
                raw = fh.read()
            protected = _password_protect_zip_bytes(raw, "naf_messaging_backup.db", password, export_id)
            log_event(current_user.id, "SYSTEM_DB_RECOVERY_EXPORTED", {"export_id": export_id, "password_protected": True, "status": "SUCCESS"})
            return send_file(io.BytesIO(protected), mimetype="application/zip", as_attachment=True, download_name=f"{export_id}_protected.zip")
    abort(404)


@bp.route("/backup/export", methods=["GET", "POST"])
@login_required
@require_roles(Role.ADMIN.value)
def export_backup():
    export_id = f"NMS-RECOVERY-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}"
    if request.method == "GET":
        return render_template(
            "protected_export_password.html",
            export_title="Disaster Recovery ZIP",
            export_id=export_id,
            suggested_password=_suggest_export_password("REC"),
            action_url=url_for("admin.export_backup"),
            cancel_url=url_for("admin.settings"),
            hidden_fields={"export_id": export_id},
        )
    password, err = _resolve_export_password()
    if err:
        flash(err, "warning")
        return redirect(url_for("admin.export_backup"))
    try:
        backup_bytes, name = build_backup_zip_bytes(current_app, include_database=True)
        protected = _password_protect_zip_bytes(backup_bytes, name, password, export_id)
    except RuntimeError as exc:
        flash(str(exc), "warning")
        return redirect(url_for("admin.settings"))
    log_event(current_user.id, "SYSTEM_BACKUP_EXPORTED", {"file": name, "export_id": export_id, "password_protected": True, "status": "SUCCESS"})
    return send_file(io.BytesIO(protected), mimetype="application/zip", as_attachment=True, download_name=f"{export_id}_protected.zip")


@bp.get("/backup/list")
@login_required
@require_roles(Role.ADMIN.value)
def list_backups():
    folder = backup_folder(current_app)
    rows = []
    for file_name in sorted(os.listdir(folder), reverse=True):
        path = os.path.join(folder, file_name)
        if os.path.isfile(path) and file_name.endswith('.zip'):
            rows.append(f"{datetime.utcfromtimestamp(os.path.getmtime(path)).strftime('%Y-%m-%d %H:%M:%S')} UTC | {file_name} | {round(os.path.getsize(path)/(1024*1024), 2)} MB")
    return render_template("admin/system_logs.html", lines=rows or ["No backup files generated yet."])
